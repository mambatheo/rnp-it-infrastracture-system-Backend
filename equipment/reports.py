import io
import os
from django.conf import settings
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image as RLImage,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from .models import Equipment, EquipmentCategory, EquipmentStatus, Stock, Unit, Region, DPU
from dotenv import load_dotenv

load_dotenv()

_FONTS_DIR = os.getenv("PDF_FONT_DIR", "./static/fonts")

try:
    pdfmetrics.registerFont(TTFont("Tahoma",      os.path.join(_FONTS_DIR, "tahoma.ttf")))
    pdfmetrics.registerFont(TTFont("Tahoma-Bold", os.path.join(_FONTS_DIR, "tahomabd.ttf")))
    _PDF_FONT      = "Tahoma"
    _PDF_FONT_BOLD = "Tahoma-Bold"
except Exception as e:
    print(f"[WARN] Tahoma font not loaded ({e}), falling back to Helvetica")
    _PDF_FONT      = "Helvetica"
    _PDF_FONT_BOLD = "Helvetica-Bold"


# ─────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────

SYSTEM_NAME     = os.getenv("SYSTEM_NAME")
LOGO_PATH       = os.path.join(settings.BASE_DIR, "static", "images", "rnp_logo.png")
REPORT_PASSWORD = os.getenv("EXCEL_REPORT_PASSWORD")


def get_equipment_types():
    return (
        Equipment.objects.values_list("equipment_type", flat=True)
        .distinct()
        .order_by("equipment_type")
    )

# ─────────────────────────────────────────
# COLOURS
# ─────────────────────────────────────────

DARK_BLUE = "1F4E79"
ALT_ROW   = "EBF3FB"
WHITE     = "FFFFFF"
ACCENT    = "2E75B6"

# ── Pre-built openpyxl style singletons (reuse to avoid per-cell allocation) ──
_THIN_BORDER_SIDE  = Side(style="thin", color="CCCCCC")
_WHITE_BORDER_SIDE = Side(style="thin", color="FFFFFF")

_DATA_BORDER     = Border(left=_THIN_BORDER_SIDE,  right=_THIN_BORDER_SIDE,
                          top=_THIN_BORDER_SIDE,   bottom=_THIN_BORDER_SIDE)
_HEADER_BORDER   = Border(left=_WHITE_BORDER_SIDE, right=_WHITE_BORDER_SIDE,
                          top=_WHITE_BORDER_SIDE,  bottom=_WHITE_BORDER_SIDE)

_FILL_WHITE      = PatternFill(start_color=WHITE,    end_color=WHITE,    fill_type="solid")
_FILL_ALT        = PatternFill(start_color=ALT_ROW,  end_color=ALT_ROW,  fill_type="solid")
_FILL_DARK_BLUE  = PatternFill(start_color=DARK_BLUE,end_color=DARK_BLUE,fill_type="solid")
_FILL_ACCENT     = PatternFill(start_color=ACCENT,   end_color=ACCENT,   fill_type="solid")

_ALIGN_CENTER         = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_CENTER_NO_WRAP = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT          = Alignment(horizontal="right",  vertical="center")

_FONT_DATA       = Font(name="Tahoma", size=11)
_FONT_HEADER     = Font(name="Tahoma", color=WHITE, bold=True, size=11)
_FONT_GRAND      = Font(name="Tahoma", bold=True, color=WHITE, size=12)
_FONT_BANNER     = Font(name="Tahoma", bold=True, color="FFFFFF", size=13)
_FONT_SECTION    = Font(name="Tahoma", bold=True, color="FFFFFF", size=12)

# Bug 1 fix: pre-build Protection singleton — avoids creating one object per cell in _protect_sheet
_LOCKED_PROTECTION = Protection(locked=True)

PDF_DARK   = colors.HexColor("#1F4E79")
PDF_ALT    = colors.HexColor("#EBF3FB")
PDF_ACCENT = colors.HexColor("#2E75B6")
PDF_GREY   = colors.HexColor("#CCCCCC")

# ─────────────────────────────────────────
# FIELDS
# ─────────────────────────────────────────

# FIX: added missing comma between "Model" and "Serial Number"
BASIC_FIELDS = [
    "S/N", "Brand", "Model", "Serial Number", "Marking Code",
    "Location", "Status", "Age",
]

COL_WIDTHS = [
    1.3*cm,   # S/N
    2.0*cm,   # Brand
    4.5*cm,   # Model
    4.5*cm,   # Serial Number
    4.5*cm,   # Marking Code
    7.0*cm,   # Location
    3.0*cm,   # Status
    3.0*cm,   # Age
]

# Landscape A4 content width: 297mm − 10mm left − 10mm right = 277mm
PAGE_CONTENT_WIDTH = 27.7 * cm

# Maps header label → PDF column width (used for dynamic filtering)
_FIELD_WIDTH = dict(zip(BASIC_FIELDS, COL_WIDTHS))


def _filter_empty_cols(headers, rows, widths=None):

    if not rows:
        return headers, rows, widths

    keep = [
        col_idx
        for col_idx in range(len(headers))
        if any(str(row[col_idx]).strip() not in ("—", "-", "") for row in rows)
    ]

    f_headers = [headers[i] for i in keep]
    f_rows    = [[row[i] for i in keep] for row in rows]
    f_widths  = [widths[i] for i in keep] if widths else None
    return f_headers, f_rows, f_widths


def _scale_to_page(widths, page_width=PAGE_CONTENT_WIDTH):
    """
    Proportionally scale column widths so they sum exactly to page_width.
    """
    total = sum(widths) if widths else 0
    if not total:
        return widths
    factor = page_width / total
    return [w * factor for w in widths]


# ── Paragraph styles used inside PDF table cells ───────────────────────────
# Built once at module level to avoid ReportLab name-collision KeyErrors on
# repeated requests (Bugs 4/5/6) and to avoid per-call object allocation.
_PDF_STYLES = getSampleStyleSheet()

_CELL_STYLE = ParagraphStyle(
    "CellNormal",
    fontName=_PDF_FONT,
    fontSize=11,
    leading=13,
    wordWrap="LTR",
    alignment=1,
)
_HEADER_CELL_STYLE = ParagraphStyle(
    "CellHeader",
    fontName=_PDF_FONT_BOLD,
    fontSize=11,
    leading=13,
    wordWrap="LTR",
    textColor=colors.white,
    alignment=1,
)

# PDF heading / body styles — shared across all report generators
_STYLE_PDF_SYSTEM = ParagraphStyle(
    "PDFSystemName", parent=_PDF_STYLES["Normal"],
    fontSize=14, fontName=_PDF_FONT_BOLD,
    textColor=PDF_DARK, alignment=2,
)
_STYLE_PDF_REPORT = ParagraphStyle(
    "PDFReportTitle", parent=_PDF_STYLES["Normal"],
    fontSize=12, fontName=_PDF_FONT,
    textColor=colors.HexColor("#444444"), alignment=2,
)
_STYLE_PDF_DATE = ParagraphStyle(
    "PDFDate", parent=_PDF_STYLES["Normal"],
    fontSize=10, fontName=_PDF_FONT,
    textColor=colors.grey, alignment=2,
)
_STYLE_PDF_HEADING = ParagraphStyle(
    "PDFHeading", parent=_PDF_STYLES["Heading2"],
    textColor=PDF_DARK, fontName=_PDF_FONT_BOLD,
)
_STYLE_PDF_SMALL = ParagraphStyle(
    "PDFSmall", parent=_PDF_STYLES["Normal"],
    fontSize=10, fontName=_PDF_FONT, textColor=colors.grey,
)
_STYLE_SEC_HEAD = ParagraphStyle(
    "PDFSecHead", parent=_PDF_STYLES["Normal"],
    fontSize=12, fontName=_PDF_FONT_BOLD,
    textColor=colors.white,
    backColor=colors.HexColor("#2E4DA0"),
    spaceAfter=0, spaceBefore=4,
    leftIndent=4,
)
_STYLE_UNIT_HEAD = ParagraphStyle(
    "PDFUnitHead", parent=_PDF_STYLES["Heading1"],
    textColor=colors.HexColor("#003580"),
    fontSize=11, fontName=_PDF_FONT_BOLD,
)
_STYLE_REGION_HEAD = ParagraphStyle(
    "PDFRegionHead", parent=_PDF_STYLES["Heading1"],
    textColor=colors.HexColor("#003580"),
    fontSize=11, fontName=_PDF_FONT_BOLD,
)
_STYLE_DPU_HEAD = ParagraphStyle(
    "PDFDPUHead", parent=_PDF_STYLES["Heading1"],
    textColor=colors.HexColor("#003580"),
    fontSize=11, fontName=_PDF_FONT_BOLD,
)


def _wrap_rows(header_row, data_rows):
    """Convert plain-string table data into Paragraph-wrapped lists."""
    wrapped_header = [Paragraph(str(h), _HEADER_CELL_STYLE) for h in header_row]
    wrapped_data   = [
        [Paragraph(str(cell), _CELL_STYLE) for cell in row]
        for row in data_rows
    ]
    return wrapped_header, wrapped_data


# ─────────────────────────────────────────
# SHEET PROTECTION
# ─────────────────────────────────────────

def _protect_sheet(ws):
    # Reuse singleton — avoids creating thousands of Protection() objects per sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = _LOCKED_PROTECTION
    ws.protection.sheet                = True
    ws.protection.password             = REPORT_PASSWORD
    ws.protection.enable()
    ws.protection.sort                 = False
    ws.protection.autoFilter           = False
    ws.protection.selectLockedCells    = False
    ws.protection.selectUnlockedCells  = False

# ─────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────

STATUS_KEYS = [
    ("active",       "Active"),
    ("inactive",     "Inactive"),
    ("under_repair", "Under Repair"),
    ("retired",      "Retired"),
    ("damaged",      "Damaged"),
    ("new",          "New"),
]


def get_summary():
    summary      = {}
    grand_total  = 0

    for eq_type in get_equipment_types():
        qs    = Equipment.objects.filter(equipment_type=eq_type)
        count = qs.count()
        grand_total += count
        stats = {"total": count}
        for key, label in STATUS_KEYS:
            stats[key] = qs.filter(status__name__iexact=label).count()
        summary[eq_type] = stats

    summary["__grand_total__"] = grand_total
    return summary


def get_basic_rows(queryset=None, equipment_type=None):
    if queryset is None:
        qs = Equipment.objects.select_related(
            "brand", "status",
            "region", "region__region_office",
            "dpu", "dpu__dpu_office",
            "station",
            "unit", "directorate", "department", "office",
        )
        if equipment_type:
            qs = qs.filter(equipment_type=equipment_type)
    else:
        qs = queryset

    rows = []
    for sn, obj in enumerate(qs, start=1):
        parts = []
        if obj.office:
            parts.append(str(obj.office))
        elif obj.unit:
            parts.append(str(obj.unit))
        elif obj.dpu:
            parts.append(str(obj.dpu))
            if obj.dpu.dpu_office:
                parts.append(str(obj.dpu.dpu_office))
        elif obj.region:
            parts.append(str(obj.region))
            if obj.region.region_office:
                parts.append(str(obj.region.region_office))
        location = ", ".join(parts) if parts else "—"

        # FIX: rows now include Brand and Model to match BASIC_FIELDS (8 columns)
        rows.append([
            str(sn),
            str(obj.brand) if obj.brand else "—",  # Brand
            obj.model              or "—",           # Model
            obj.serial_number      or "—",           # Serial Number
            obj.marking_code       or "—",           # Marking Code
            location,                                # Location
            str(obj.status) if obj.status else "—", # Status
            obj.age_since_deployed or "—",           # Age
        ])
    return rows


# ─────────────────────────────────────────
# EXCEL CELL HELPERS
# ─────────────────────────────────────────

def _header_cell(cell):
    cell.font      = _FONT_HEADER
    cell.fill      = _FILL_DARK_BLUE
    cell.alignment = _ALIGN_CENTER
    cell.border    = _HEADER_BORDER


def _data_cell(cell, alt=False):
    cell.fill      = _FILL_ALT if alt else _FILL_WHITE
    cell.alignment = _ALIGN_CENTER
    cell.border    = _DATA_BORDER
    cell.font      = _FONT_DATA


def _excel_header(ws, report_title, col_count):
    ws.sheet_view.showGridLines      = False
    ws.page_setup.orientation        = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage          = True
    ws.page_setup.fitToWidth         = 1
    ws.row_dimensions[1].height      = 80
    ws.row_dimensions[2].height      = 22
    ws.row_dimensions[3].height      = 18

    if os.path.exists(LOGO_PATH):
        logo        = XLImage(LOGO_PATH)
        logo.width  = 140
        logo.height = 112
        ws.add_image(logo, "A1")

    mid = max(col_count // 2, 1)

    ws.merge_cells(start_row=1, start_column=mid + 1, end_row=1, end_column=col_count)
    sys_cell           = ws.cell(row=1, column=mid + 1, value=SYSTEM_NAME)
    sys_cell.font      = Font(name="Tahoma", size=14, bold=True, color=DARK_BLUE)
    sys_cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(start_row=2, start_column=mid + 1, end_row=2, end_column=col_count)
    title_cell           = ws.cell(row=2, column=mid + 1, value=report_title)
    title_cell.font      = Font(name="Tahoma", size=12, color=ACCENT)
    title_cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(start_row=3, start_column=mid + 1, end_row=3, end_column=col_count)
    date_cell = ws.cell(
        row=3, column=mid + 1,
        value=f"Generated: {timezone.now().strftime('%d %B %Y, %H:%M')}",
    )
    date_cell.font      = Font(name="Tahoma", size=10, italic=True, color="666666")
    date_cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.append([])  # row 4 spacer


def _write_equipment_sheet(ws, title, equipment_type=None, queryset=None):
    rows = get_basic_rows(queryset=queryset, equipment_type=equipment_type)
    headers, rows, _ = _filter_empty_cols(BASIC_FIELDS, rows)

    _excel_header(ws, f"Equipment Report — {title}", len(headers))

    ws.append(headers)
    for cell in ws[ws.max_row]:
        _header_cell(cell)
    ws.row_dimensions[ws.max_row].height = 20

    if not rows:
        ws.append(["No data available"] + [""] * (len(headers) - 1))
    else:
        for i, row in enumerate(rows):
            ws.append(row)
            for cell in ws[ws.max_row]:
                _data_cell(cell, alt=(i % 2 == 0))
            ws.row_dimensions[ws.max_row].height = 16

    for col in ws.columns:
        first = col[0]
        if isinstance(first, MergedCell):
            continue
        max_len = max(
            (len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)),
            default=10,
        )
        ws.column_dimensions[first.column_letter].width = min(max_len + 4, 35)

    _protect_sheet(ws)

# ─────────────────────────────────────────
# PDF HELPERS
# ─────────────────────────────────────────

TABLE_STYLE = TableStyle([
    ("BACKGROUND",   (0, 0),  (-1, 0),  PDF_DARK),
    ("TEXTCOLOR",    (0, 0),  (-1, 0),  colors.white),
    ("FONTNAME",     (0, 0),  (-1, 0),  _PDF_FONT_BOLD),
    ("FONTNAME",     (0, 1),  (-1, -1), _PDF_FONT),
    ("FONTSIZE",     (0, 0),  (-1, -1), 11),
    ("ALIGN",        (0, 0),  (-1, -1), "CENTER"),
    ("VALIGN",       (0, 0),  (-1, -1), "MIDDLE"),
    ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, PDF_ALT]),
    ("GRID",         (0, 0),  (-1, -1), 0.4, PDF_GREY),
    ("ROWHEIGHT",    (0, 0),  (-1, -1), 18),
    ("TOPPADDING",   (0, 0),  (-1, -1), 4),
    ("BOTTOMPADDING",(0, 0),  (-1, -1), 4),
])


def _pdf_header(report_title):
    logo_cell = (
        RLImage(LOGO_PATH, width=2.5*cm, height=2.5*cm)
        if os.path.exists(LOGO_PATH)
        else Paragraph("[ Logo ]", _PDF_STYLES["Normal"])
    )

    right_content = [
        Paragraph(SYSTEM_NAME, _STYLE_PDF_SYSTEM),
        Spacer(1, 0.1*cm),
        Paragraph(report_title, _STYLE_PDF_REPORT),
        Spacer(1, 0.1*cm),
        Paragraph(f"Generated: {timezone.now().strftime('%d %B %Y at %H:%M')}", _STYLE_PDF_DATE),
    ]

    header_table = Table([[logo_cell, right_content]], colWidths=[4*cm, None], hAlign="LEFT")
    header_table.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",        (0, 0), (0, 0),   "LEFT"),
        ("ALIGN",        (1, 0), (1, 0),   "RIGHT"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    return header_table


def _pdf_summary_table(summary):
    headers = ["Equipment Type", "Total", "Active", "Inactive", "Under Repair", "Retired", "Damaged", "New"]
    data    = [headers]

    for eq_type, stats in summary.items():
        if eq_type == "__grand_total__":
            continue
        data.append([
            eq_type,
            stats["total"],   stats["active"],   stats["inactive"],
            stats["under_repair"], stats["retired"], stats["damaged"], stats["new"],
        ])

    data.append(["GRAND TOTAL", summary["__grand_total__"], "", "", "", "", "", ""])

    w_header, w_data = _wrap_rows(data[0], data[1:])
    t = Table([w_header] + w_data, repeatRows=1, hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  PDF_DARK),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0),  (-1, 0),  _PDF_FONT_BOLD),
        ("FONTNAME",      (0, 1),  (-1, -1), _PDF_FONT),
        ("FONTSIZE",      (0, 0),  (-1, -1), 12),
        ("ALIGN",         (0, 0),  (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0, 1),  (-1, -2), [colors.white, PDF_ALT]),
        ("BACKGROUND",    (0, -1), (-1, -1), PDF_ACCENT),
        ("TEXTCOLOR",     (0, -1), (-1, -1), colors.white),
        ("FONTNAME",      (0, -1), (-1, -1), _PDF_FONT_BOLD),
        ("GRID",          (0, 0),  (-1, -1), 0.5, PDF_GREY),
        ("ROWHEIGHT",     (0, 0),  (-1, -1), 20),
        ("TOPPADDING",    (0, 0),  (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
    ]))
    return t


def _pdf_equipment_section(equipment_type):
    elements = []
    elements.append(HRFlowable(width="100%", thickness=1, color=PDF_GREY))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(equipment_type, _STYLE_PDF_HEADING))

    rows = get_basic_rows(equipment_type=equipment_type)
    if not rows:
        elements.append(Paragraph("No data available.", _STYLE_PDF_SMALL))
        elements.append(Spacer(1, 0.4*cm))
        return elements

    headers, rows, col_widths = _filter_empty_cols(
        BASIC_FIELDS, rows,
        widths=[_FIELD_WIDTH[h] for h in BASIC_FIELDS],
    )

    w_header, w_data = _wrap_rows(headers, rows)
    t = Table([w_header] + w_data, repeatRows=1, hAlign="LEFT", colWidths=_scale_to_page(col_widths))
    t.setStyle(TABLE_STYLE)
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))
    return elements


def _build_pdf(elements):
    output = io.BytesIO()
    doc    = SimpleDocTemplate(
        output,
        pagesize     = landscape(A4),
        rightMargin  = 1*cm,  leftMargin  = 1*cm,
        topMargin    = 1.5*cm, bottomMargin = 1.5*cm,
    )
    doc.build(elements)
    output.seek(0)
    return output

# ─────────────────────────────────────────
# EXCEL — ALL EQUIPMENT IN ONE WORKBOOK
# ─────────────────────────────────────────

def generate_excel_all():
    wb         = Workbook()
    summary    = get_summary()
    eq_types   = list(get_equipment_types())

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary       = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_view.showGridLines = False
    _excel_header(ws_summary, "Equipment Summary Report", 8)

    headers = ["Equipment Type", "Total", "Active", "Inactive", "Under Repair", "Retired", "Damaged", "New"]
    ws_summary.append(headers)
    for cell in ws_summary[ws_summary.max_row]:
        _header_cell(cell)
    ws_summary.row_dimensions[ws_summary.max_row].height = 22

    for i, (eq_type, stats) in enumerate(summary.items()):
        if eq_type == "__grand_total__":
            continue
        ws_summary.append([
            eq_type,
            stats["total"],   stats["active"],   stats["inactive"],
            stats["under_repair"], stats["retired"], stats["damaged"], stats["new"],
        ])
        for cell in ws_summary[ws_summary.max_row]:
            _data_cell(cell, alt=(i % 2 == 0))
        ws_summary.row_dimensions[ws_summary.max_row].height = 18

    ws_summary.append(["GRAND TOTAL", summary["__grand_total__"], "", "", "", "", "", ""])
    for cell in ws_summary[ws_summary.max_row]:
        cell.font      = _FONT_GRAND
        cell.fill      = _FILL_ACCENT
        cell.alignment = _ALIGN_CENTER_NO_WRAP
    ws_summary.row_dimensions[ws_summary.max_row].height = 22

    for i, w in enumerate([28, 8, 8, 10, 14, 10, 10, 8], 1):
        ws_summary.column_dimensions[chr(64 + i)].width = w
    _protect_sheet(ws_summary)

    # ── One sheet per equipment type ──────────────────────────────────────────
    for eq_type in eq_types:
        ws = wb.create_sheet(title=eq_type[:31])
        _write_equipment_sheet(ws, eq_type, equipment_type=eq_type)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────
# EXCEL — SINGLE EQUIPMENT TYPE
# ─────────────────────────────────────────

def generate_excel_by_type(equipment_type):
    wb    = Workbook()
    ws    = wb.active
    ws.title = equipment_type[:31]
    _write_equipment_sheet(ws, equipment_type, equipment_type=equipment_type)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────
# PDF — ALL EQUIPMENT IN ONE FILE
# ─────────────────────────────────────────

def generate_pdf_all():
    elements = [
        _pdf_header("Full Equipment Report"),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=2, color=PDF_ACCENT),
        Spacer(1, 0.5*cm),
        Paragraph("Summary", _STYLE_PDF_HEADING),
        Spacer(1, 0.2*cm),
        _pdf_summary_table(get_summary()),
        Spacer(1, 0.8*cm),
    ]

    for eq_type in get_equipment_types():
        elements.extend(_pdf_equipment_section(eq_type))

    return _build_pdf(elements)


# ─────────────────────────────────────────
# PDF — SINGLE EQUIPMENT TYPE
# ─────────────────────────────────────────

def generate_pdf_by_type(equipment_type):
    elements = [
        _pdf_header(f"{equipment_type} Report"),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=2, color=PDF_ACCENT),
        Spacer(1, 0.5*cm),
    ]
    elements.extend(_pdf_equipment_section(equipment_type))
    return _build_pdf(elements)


# ═════════════════════════════════════════════════════════════════
#  STOCK REPORTS
# ═════════════════════════════════════════════════════════════════

STOCK_FIELDS = [
    "S/N", "Serial Number", "Marking Code",
    "Brand", "Model", "Status", "Condition", "Date Added",
]

STOCK_COL_WIDTHS = [
    1.2*cm,   # S/N
    4.0*cm,   # Serial Number
    4.0*cm,   # Marking Code
    3.0*cm,   # Brand
    3.0*cm,   # Model
    3.0*cm,   # Status
    3.0*cm,   # Condition
    3.5*cm,   # Date Added
]

_STOCK_FIELD_WIDTH = dict(zip(STOCK_FIELDS, STOCK_COL_WIDTHS))


def _get_stock_qs(equipment_type=None):
    qs = Stock.objects.select_related(
        "equipment__brand", "equipment__status",
        "added_by",
    ).order_by("equipment__equipment_type", "equipment__name")
    if equipment_type:
        qs = qs.filter(equipment__equipment_type=equipment_type)
    return qs


def _stock_row(sn, s):
    eq = s.equipment
    return [
        str(sn),
        eq.serial_number         or "—",
        eq.marking_code          or "—",
        str(eq.brand)  if eq.brand  else "—",
        eq.model                 or "—",
        str(eq.status) if eq.status else "—",
        s.condition,
        str(s.date_added),
    ]


# ─── Excel helpers ────────────────────────────────────────────────

def _write_stock_sheet(ws, title, equipment_type=None):
    rows = [_stock_row(sn, s) for sn, s in enumerate(_get_stock_qs(equipment_type), start=1)]
    headers, rows, _ = _filter_empty_cols(STOCK_FIELDS, rows)

    _excel_header(ws, f"Stock Report — {title}", len(headers))

    ws.append(headers)
    for cell in ws[ws.max_row]:
        _header_cell(cell)
    ws.row_dimensions[ws.max_row].height = 20

    if not rows:
        ws.append(["No stock items available"] + [""] * (len(headers) - 1))
    else:
        for i, row in enumerate(rows):
            ws.append(row)
            for cell in ws[ws.max_row]:
                _data_cell(cell, alt=(i % 2 == 0))
            ws.row_dimensions[ws.max_row].height = 16

    for col in ws.columns:
        first = col[0]
        if isinstance(first, MergedCell):
            continue
        max_len = max(
            (len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)),
            default=10,
        )
        ws.column_dimensions[first.column_letter].width = min(max_len + 4, 40)

    _protect_sheet(ws)


def generate_stock_excel_all():
    wb = Workbook()

    ws_sum       = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _excel_header(ws_sum, "Stock Summary", 3)
    ws_sum.append(["Equipment Type", "Items in Stock", ""])
    for cell in ws_sum[ws_sum.max_row]:
        _header_cell(cell)

    types_in_stock = (
        Stock.objects.values_list("equipment__equipment_type", flat=True)
        .distinct().order_by("equipment__equipment_type")
    )
    for i, eq_type in enumerate(types_in_stock):
        count = Stock.objects.filter(equipment__equipment_type=eq_type).count()
        ws_sum.append([eq_type, count, ""])
        for cell in ws_sum[ws_sum.max_row]:
            _data_cell(cell, alt=(i % 2 == 0))

    ws_sum.append(["TOTAL", Stock.objects.count(), ""])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font      = _FONT_GRAND
        cell.fill      = _FILL_ACCENT
        cell.alignment = _ALIGN_CENTER_NO_WRAP

    for w, col in zip([30, 15, 10], ["A", "B", "C"]):
        ws_sum.column_dimensions[col].width = w
    _protect_sheet(ws_sum)

    for eq_type in types_in_stock:
        ws = wb.create_sheet(title=eq_type[:31])
        _write_stock_sheet(ws, eq_type, equipment_type=eq_type)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_stock_excel_by_type(equipment_type):
    wb = Workbook()
    ws = wb.active
    ws.title = equipment_type[:31]
    _write_stock_sheet(ws, equipment_type, equipment_type=equipment_type)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── PDF helpers ──────────────────────────────────────────────────

def _pdf_stock_section(equipment_type):
    elements = [
        HRFlowable(width="100%", thickness=1, color=PDF_GREY),
        Spacer(1, 0.3*cm),
        Paragraph(equipment_type, _STYLE_PDF_HEADING),
    ]
    rows = [_stock_row(sn, s) for sn, s in enumerate(_get_stock_qs(equipment_type), start=1)]
    if not rows:
        elements.append(Paragraph("No stock items available.", _STYLE_PDF_SMALL))
        elements.append(Spacer(1, 0.4*cm))
        return elements

    _stock_col_widths = [_STOCK_FIELD_WIDTH[h] for h in STOCK_FIELDS]
    headers, rows, col_widths = _filter_empty_cols(STOCK_FIELDS, rows, widths=_stock_col_widths)

    w_header, w_data = _wrap_rows(headers, rows)
    t = Table([w_header] + w_data, repeatRows=1, hAlign="LEFT", colWidths=_scale_to_page(col_widths))
    t.setStyle(TABLE_STYLE)
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))
    return elements


def generate_stock_pdf_all():
    types_in_stock = (
        Stock.objects.values_list("equipment__equipment_type", flat=True)
        .distinct().order_by("equipment__equipment_type")
    )
    summary_data = [["Equipment Type", "Items in Stock"]]
    for eq_type in types_in_stock:
        summary_data.append([eq_type, Stock.objects.filter(equipment__equipment_type=eq_type).count()])
    summary_data.append(["TOTAL", Stock.objects.count()])

    sw_header, sw_data = _wrap_rows(summary_data[0], summary_data[1:])
    summary_tbl = Table([sw_header] + sw_data, repeatRows=1, hAlign="LEFT")
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  PDF_DARK),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0),  (-1, 0),  _PDF_FONT_BOLD),
        ("FONTNAME",      (0, 1),  (-1, -1), _PDF_FONT),
        ("FONTSIZE",      (0, 0),  (-1, -1), 12),
        ("ALIGN",         (0, 0),  (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0, 1),  (-1, -2), [colors.white, PDF_ALT]),
        ("BACKGROUND",    (0, -1), (-1, -1), PDF_ACCENT),
        ("TEXTCOLOR",     (0, -1), (-1, -1), colors.white),
        ("FONTNAME",      (0, -1), (-1, -1), _PDF_FONT_BOLD),
        ("GRID",          (0, 0),  (-1, -1), 0.5, PDF_GREY),
        ("ROWHEIGHT",     (0, 0),  (-1, -1), 20),
        ("TOPPADDING",    (0, 0),  (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
    ]))

    elements = [
        _pdf_header("Full Stock Report"),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=2, color=PDF_ACCENT),
        Spacer(1, 0.5*cm),
        Paragraph("Summary", _STYLE_PDF_HEADING),
        Spacer(1, 0.2*cm),
        summary_tbl,
        Spacer(1, 0.8*cm),
    ]
    for eq_type in types_in_stock:
        elements.extend(_pdf_stock_section(eq_type))

    return _build_pdf(elements)


def generate_stock_pdf_by_type(equipment_type):
    elements = [
        _pdf_header(f"{equipment_type} Stock Report"),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=2, color=PDF_ACCENT),
        Spacer(1, 0.5*cm),
    ]
    elements.extend(_pdf_stock_section(equipment_type))
    return _build_pdf(elements)


# ═════════════════════════════════════════════════════════════════
#  UNIT / REGION / DPU REPORTS  — shared fields
# ═════════════════════════════════════════════════════════════════



UNIT_FIELDS = [
    "S/N", "Serial Number", "Marking Code",
    "Location", "Status", "Age",
]

UNIT_COL_WIDTHS = [
    1.2*cm,   # S/N
    4.5*cm,   # Serial Number
    4.5*cm,   # Marking Code
    5.0*cm,   # Location
    4.0*cm,   # Status
    4.0*cm,   # Age
]

_UNIT_FIELD_WIDTH = dict(zip(UNIT_FIELDS, UNIT_COL_WIDTHS))

_TYPE_ORDER = [
    "Desktop", "Laptop", "Server",
    "Printer", "Network Device", "Projector",
    "TV Screen", "Decoder", "Telephone",
    "External Storage", "Peripheral", "UPS",
]

_TYPE_LABELS = {
    "Desktop":          "DESKTOPS",
    "Laptop":           "LAPTOPS",
    "Server":           "SERVERS",
    "Printer":          "PRINTERS & SCANNERS",
    "Network Device":   "NETWORK / ACCESS POINTS",
    "Projector":        "PROJECTORS",
    "TV Screen":        "TV SCREENS",
    "Decoder":          "DECODERS",
    "Telephone":        "TELEPHONES",
    "External Storage": "EXTERNAL STORAGE",
    "Peripheral":       "PERIPHERALS",
    "UPS":              "UPS / BATTERIES",
}


def _unit_device_rows(unit_qs, equipment_type):
    """Return numbered data rows for one equipment_type within a group queryset."""
    qs = unit_qs.filter(equipment_type=equipment_type).order_by("name")
    rows = []
    for sn, obj in enumerate(qs, start=1):
        location = (
            str(obj.office) if obj.office else
            str(obj.unit)   if obj.unit   else
            str(obj.region) if obj.region else "—"
        )
        rows.append([
            str(sn),
            obj.serial_number                    or "—",
            obj.marking_code                     or "—",
            location,
            str(obj.status) if obj.status        else "—",
            obj.age_since_deployed               or "—",
        ])
    return rows


def _base_unit_qs():
    return Equipment.objects.select_related("brand", "status", "office", "unit", "region")


def _base_region_qs():
    return Equipment.objects.select_related("brand", "status", "office", "unit", "region")


def _base_dpu_qs():
    return Equipment.objects.select_related("brand", "status", "office", "unit", "region")


# ─── Excel helpers ────────────────────────────────────────────────────────────

_SECTION_FILL      = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_SECTION_ALT_FILL  = PatternFill(start_color="2E4DA0", end_color="2E4DA0", fill_type="solid")
_UNIT_HEADER_FILL  = PatternFill(start_color="003580", end_color="003580", fill_type="solid")


def _write_unit_sheet(ws, unit_name, unit_qs):
    n_cols = len(UNIT_FIELDS)
    col_letters = [chr(65 + i) for i in range(n_cols)]

    ws.sheet_view.showGridLines      = False
    ws.page_setup.orientation        = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage          = True
    ws.page_setup.fitToWidth         = 1

    # Unit name banner
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=n_cols)
    banner = ws.cell(row=1, column=1, value=unit_name.upper())
    banner.font      = _FONT_BANNER
    banner.fill      = _UNIT_HEADER_FILL
    banner.alignment = _ALIGN_CENTER_NO_WRAP
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 10

    row_num     = 3
    section_num = 0

    for eq_type in _TYPE_ORDER:
        rows = _unit_device_rows(unit_qs, eq_type)
        if not rows:
            continue
        section_num += 1
        label = f"{section_num}. {_TYPE_LABELS.get(eq_type, eq_type.upper())}"

        # Section heading row
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=n_cols
        )
        sec_cell = ws.cell(row=row_num, column=1, value=label)
        sec_cell.font      = _FONT_SECTION
        sec_cell.fill      = _SECTION_ALT_FILL
        sec_cell.alignment = _ALIGN_CENTER_NO_WRAP
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        # Column headers
        for col_idx, header in enumerate(UNIT_FIELDS, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            _header_cell(cell)
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        # Data rows
        for i, row in enumerate(rows):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                _data_cell(cell, alt=(i % 2 == 0))
            ws.row_dimensions[row_num].height = 15
            row_num += 1

        row_num += 1  # spacer between sections

    # Column widths
    for letter, width in zip(col_letters, UNIT_COL_WIDTHS):
        ws.column_dimensions[letter].width = width / cm * 2.54 * 2.2

    _protect_sheet(ws)


# ═════════════════════════════════════════════════════════════════
#  UNIT REPORTS
# ═════════════════════════════════════════════════════════════════

def generate_unit_excel_all():
    wb = Workbook()

    ws_sum        = wb.active
    ws_sum.title  = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _excel_header(ws_sum, "Equipment by Organisational Unit", 3)
    ws_sum.append(["Unit", "Equipment Count", ""])
    for cell in ws_sum[ws_sum.max_row]:
        _header_cell(cell)

    units = Unit.objects.order_by("name")
    base  = _base_unit_qs()
    grand = 0
    for i, unit in enumerate(units):
        qs    = base.filter(unit=unit)
        count = qs.count()
        if count == 0:
            continue
        grand += count
        ws_sum.append([unit.name, count, ""])
        for cell in ws_sum[ws_sum.max_row]:
            _data_cell(cell, alt=(i % 2 == 0))
    ws_sum.append(["TOTAL", grand, ""])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font      = _FONT_GRAND
        cell.fill      = _FILL_ACCENT
        cell.alignment = _ALIGN_CENTER_NO_WRAP
    for w, col in zip([30, 15, 10], ["A", "B", "C"]):
        ws_sum.column_dimensions[col].width = w
    _protect_sheet(ws_sum)

    for unit in units:
        qs = base.filter(unit=unit)
        if not qs.exists():
            continue
        ws = wb.create_sheet(title=unit.name[:31])
        _write_unit_sheet(ws, unit.name, qs)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_unit_excel_by_unit(unit_id):
    unit = Unit.objects.get(pk=unit_id)
    qs   = _base_unit_qs().filter(unit=unit)
    wb   = Workbook()
    ws   = wb.active
    ws.title = unit.name[:31]
    _write_unit_sheet(ws, unit.name, qs)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── PDF helpers ──────────────────────────────────────────────────────────────

_UNIT_SCALED_WIDTHS = _scale_to_page(UNIT_COL_WIDTHS)


def _pdf_unit_section(eq_type, section_num, unit_qs):
    label = f"{section_num}. {_TYPE_LABELS.get(eq_type, eq_type.upper())}"
    rows = _unit_device_rows(unit_qs, eq_type)
    if not rows:
        return []

    w_header, w_data = _wrap_rows(UNIT_FIELDS, rows)
    elements = [Paragraph(label, _STYLE_SEC_HEAD)]
    t = Table([w_header] + w_data, repeatRows=1, hAlign="LEFT", colWidths=_UNIT_SCALED_WIDTHS)
    t.setStyle(TABLE_STYLE)
    elements.append(t)
    elements.append(Spacer(1, 0.4*cm))
    return elements


def _pdf_unit_block(unit, base_qs):
    qs = base_qs.filter(unit=unit)
    if not qs.exists():
        return []

    elements = [
        HRFlowable(width="100%", thickness=2, color=colors.HexColor("#003580")),
        Spacer(1, 0.2*cm),
        Paragraph(unit.name.upper(), _STYLE_UNIT_HEAD),
        Spacer(1, 0.2*cm),
    ]
    section_num = 0
    for eq_type in _TYPE_ORDER:
        if qs.filter(equipment_type=eq_type).exists():
            section_num += 1
            elements.extend(_pdf_unit_section(eq_type, section_num, qs))

    elements.append(Spacer(1, 0.6*cm))
    return elements


def generate_unit_pdf_all():
    elements = [
        _pdf_header("Equipment Report by Organisational Unit"),
        Spacer(1, 0.4*cm),
    ]
    base  = _base_unit_qs()
    units = Unit.objects.order_by("name")
    for unit in units:
        elements.extend(_pdf_unit_block(unit, base))
    return _build_pdf(elements)


def generate_unit_pdf_by_unit(unit_id):
    unit     = Unit.objects.get(pk=unit_id)
    elements = [
        _pdf_header(f"{unit.name.upper()} — Equipment Report"),
        Spacer(1, 0.4*cm),
    ]
    elements.extend(_pdf_unit_block(unit, _base_unit_qs()))
    return _build_pdf(elements)


# ═════════════════════════════════════════════════════════════════
#  REGION REPORTS
# ═════════════════════════════════════════════════════════════════

def _write_region_sheet(ws, region_name, region_qs):
    _write_unit_sheet(ws, region_name, region_qs)


def generate_region_excel_all():
    wb = Workbook()

    ws_sum        = wb.active
    ws_sum.title  = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _excel_header(ws_sum, "Equipment by Region", 3)
    ws_sum.append(["Region", "Equipment Count", ""])
    for cell in ws_sum[ws_sum.max_row]:
        _header_cell(cell)

    regions = Region.objects.order_by("name")
    base    = _base_region_qs()
    grand   = 0
    for i, region in enumerate(regions):
        qs    = base.filter(region=region)
        count = qs.count()
        if count == 0:
            continue
        grand += count
        ws_sum.append([region.name, count, ""])
        for cell in ws_sum[ws_sum.max_row]:
            _data_cell(cell, alt=(i % 2 == 0))
    ws_sum.append(["TOTAL", grand, ""])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font      = _FONT_GRAND
        cell.fill      = _FILL_ACCENT
        cell.alignment = _ALIGN_CENTER_NO_WRAP
    for w, col in zip([30, 15, 10], ["A", "B", "C"]):
        ws_sum.column_dimensions[col].width = w
    _protect_sheet(ws_sum)

    for region in regions:
        qs = base.filter(region=region)
        if not qs.exists():
            continue
        ws = wb.create_sheet(title=region.name[:31])
        _write_region_sheet(ws, region.name, qs)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_region_excel_by_region(region_id):
    region = Region.objects.get(pk=region_id)
    qs     = _base_region_qs().filter(region=region)
    wb     = Workbook()
    ws     = wb.active
    ws.title = region.name[:31]
    _write_region_sheet(ws, region.name, qs)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _pdf_region_block(region, base_qs):
    qs = base_qs.filter(region=region)
    if not qs.exists():
        return []

    elements = [
        HRFlowable(width="100%", thickness=2, color=colors.HexColor("#003580")),
        Spacer(1, 0.2*cm),
        Paragraph(region.name.upper(), _STYLE_REGION_HEAD),
        Spacer(1, 0.2*cm),
    ]
    section_num = 0
    for eq_type in _TYPE_ORDER:
        if qs.filter(equipment_type=eq_type).exists():
            section_num += 1
            elements.extend(_pdf_unit_section(eq_type, section_num, qs))

    elements.append(Spacer(1, 0.6*cm))
    return elements


def generate_region_pdf_all():
    elements = [
        _pdf_header("Equipment Report by Region"),
        Spacer(1, 0.4*cm),
    ]
    base    = _base_region_qs()
    regions = Region.objects.order_by("name")
    for region in regions:
        elements.extend(_pdf_region_block(region, base))
    return _build_pdf(elements)


def generate_region_pdf_by_region(region_id):
    region   = Region.objects.get(pk=region_id)
    elements = [
        _pdf_header(f"{region.name.upper()} — Equipment Report"),
        Spacer(1, 0.4*cm),
    ]
    elements.extend(_pdf_region_block(region, _base_region_qs()))
    return _build_pdf(elements)


# ═════════════════════════════════════════════════════════════════
#  DPU REPORTS
# ═════════════════════════════════════════════════════════════════

def _write_dpu_sheet(ws, dpu_name, dpu_qs):
    _write_unit_sheet(ws, dpu_name, dpu_qs)


def generate_dpu_excel_all():
    wb = Workbook()

    ws_sum        = wb.active
    ws_sum.title  = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _excel_header(ws_sum, "Equipment by DPU", 3)
    ws_sum.append(["DPU", "Equipment Count", ""])
    for cell in ws_sum[ws_sum.max_row]:
        _header_cell(cell)

    dpus  = DPU.objects.order_by("name")
    base  = _base_dpu_qs()
    grand = 0
    for i, dpu in enumerate(dpus):
        qs    = base.filter(dpu=dpu)
        count = qs.count()
        if count == 0:
            continue
        grand += count
        ws_sum.append([dpu.name, count, ""])
        for cell in ws_sum[ws_sum.max_row]:
            _data_cell(cell, alt=(i % 2 == 0))
    ws_sum.append(["TOTAL", grand, ""])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font      = _FONT_GRAND
        cell.fill      = _FILL_ACCENT
        cell.alignment = _ALIGN_CENTER_NO_WRAP
    for w, col in zip([30, 15, 10], ["A", "B", "C"]):
        ws_sum.column_dimensions[col].width = w
    _protect_sheet(ws_sum)

    for dpu in dpus:
        qs = base.filter(dpu=dpu)
        if not qs.exists():
            continue
        ws = wb.create_sheet(title=dpu.name[:31])
        _write_dpu_sheet(ws, dpu.name, qs)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_dpu_excel_by_dpu(dpu_id):
    dpu    = DPU.objects.get(pk=dpu_id)
    qs     = _base_dpu_qs().filter(dpu=dpu)
    wb     = Workbook()
    ws     = wb.active
    ws.title = dpu.name[:31]
    _write_dpu_sheet(ws, dpu.name, qs)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _pdf_dpu_block(dpu, base_qs):
    qs = base_qs.filter(dpu=dpu)
    if not qs.exists():
        return []

    elements = [
        HRFlowable(width="100%", thickness=2, color=colors.HexColor("#003580")),
        Spacer(1, 0.2*cm),
        Paragraph(dpu.name.upper(), _STYLE_DPU_HEAD),
        Spacer(1, 0.2*cm),
    ]
    section_num = 0
    for eq_type in _TYPE_ORDER:
        if qs.filter(equipment_type=eq_type).exists():
            section_num += 1
            elements.extend(_pdf_unit_section(eq_type, section_num, qs))

    elements.append(Spacer(1, 0.6*cm))
    return elements


def generate_dpu_pdf_all():
    elements = [
        _pdf_header("Equipment Report by DPU"),
        Spacer(1, 0.4*cm),
    ]
    base = _base_dpu_qs()
    dpus = DPU.objects.order_by("name")
    for dpu in dpus:
        elements.extend(_pdf_dpu_block(dpu, base))
    return _build_pdf(elements)


def generate_dpu_pdf_by_dpu(dpu_id):
    dpu      = DPU.objects.get(pk=dpu_id)
    elements = [
        _pdf_header(f"{dpu.name.upper()} — Equipment Report"),
        Spacer(1, 0.4*cm),
    ]
    elements.extend(_pdf_dpu_block(dpu, _base_dpu_qs()))
    return _build_pdf(elements)